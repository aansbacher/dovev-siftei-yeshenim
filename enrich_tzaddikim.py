"""
enrich_tzaddikim.py
AI enrichment for tzaddikim: rewrite biography/story/torah (<=120 words),
normalize name format: הרב/רבי [name] זצ"ל

Usage:
    python enrich_tzaddikim.py --months תשרי,חשוון [--excel PATH] [--dry-run] [--resume]

Add ANTHROPIC_API_KEY to .env.local or environment.
"""

import os, re, sys, io, time, json, argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Env ───────────────────────────────────────────────────────────────────────

def load_env(path='.env.local'):
    env = {}
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1)
                    env[k.strip()] = v.strip()
    return env

_env = load_env(os.path.join(os.path.dirname(__file__), '.env.local'))
ANTHROPIC_KEY = (os.getenv('ANTHROPIC_API_KEY')
                 or _env.get('ANTHROPIC_API_KEY', ''))

# ── Name normalization ────────────────────────────────────────────────────────

_HONORIFICS = [
    'מרן הרב', 'מרן רבי', 'הגאון הגדול הרב', 'הגאון הגדול רבי',
    'הגאון הרב', 'הגאון רבי', 'הרה"ג', 'הרה"צ', 'הרה"ק',
    'האדמו"ר', 'מרן', 'הרב', 'רבנו', 'רבן', 'רבי',
]
_RABBINIC_PREFIXES = {'רבי', 'רבנו', 'רבן'}

_SUFFIX_RE = re.compile(
    r'\s*,?\s*(?:זצ"ל|זיע"א|זי"ע|זצוק"ל|ז"ל|נ"ע|עה"ש'
    r'|זכרונו לברכה|זכר צדיק לברכה)\s*$'
)
_DATE_PARENS_RE = re.compile(r'\s*\([^)]*\d{3,4}[^)]*\)')
# Also strip nikud for matching purposes
_NIKUD_RE = re.compile(r'[ְ-ׇ]')


def normalize_name(name: str) -> str:
    """Return: הרב/רבי [clean name] זצ"ל (nikud stripped from name for consistency)."""
    if not name or str(name).strip() in ('', 'nan'):
        return name

    # Strip nikud first — makes prefix matching reliable regardless of vowel marks
    clean = _NIKUD_RE.sub('', str(name).strip())

    # Decide prefix
    prefix = 'הרב'
    for p in _RABBINIC_PREFIXES:
        if clean.startswith(p):
            prefix = 'רבי'
            break

    # Strip honorific prefix (longest first)
    for h in sorted(_HONORIFICS, key=len, reverse=True):
        if clean.startswith(h):
            clean = clean[len(h):].strip()
            break

    # Strip existing suffix (זצ"ל etc.)
    clean = _SUFFIX_RE.sub('', clean).strip().strip(',').strip()

    # Strip date parentheticals like (ה'תק"ג 1743 – ה'תר"ה 1845)
    clean = _DATE_PARENS_RE.sub('', clean).strip()

    return f'{prefix} {clean} זצ"ל'


# ── AI rewriting ──────────────────────────────────────────────────────────────

_SYSTEM = """\
אתה עורך תוכן רוחני לאפליקציה יהודית-חסידית על צדיקים ורבנים.
הסגנון: חם, מכובד, נוגע ללב — כפי שמספרים סיפורי צדיקים בבית מדרש.

כללים:
- כתוב עברית תקנית עם נִיקּוּד מָלֵא.
- לא יותר מ-120 מילים.
- שמור על כל עובדה שמופיעה בטקסט המקורי.
- אל תמציא מידע שאינו בטקסט.
- כתוב בגוף שלישי.
- אל תוסיף כותרת, חתימה או הסברים — רק הטקסט עצמו.
"""

_FIELD_INTRO = {
    'biography': (
        'שכתב את הרקע הביוגרפי הבא על הרב. '
        'כתוב בסגנון חם ומכובד, עם ניקוד מלא:'
    ),
    'story': (
        'שכתב את הסיפור הבא על הרב. '
        'כתוב כסיפור חסידי נוגע ללב, עם ניקוד מלא:'
    ),
    'torah': (
        'שכתב את דברי התורה הבאים של הרב. '
        'כתוב בסגנון תורני-חסידי מרומם, עם ניקוד מלא:'
    ),
}

MODEL = 'claude-sonnet-4-5'   # Higher quality for nikud + spiritual tone
MAX_TOKENS = 500               # ~120 words with nikud takes more tokens
DELAY_SEC = 0.5                # between API calls


def rewrite_field(client, field: str, text: str, rabbi_name: str,
                  dry_run: bool) -> str:
    """Call Claude to rewrite a single field. Returns rewritten text."""
    text = str(text).strip()
    if not text or text == 'nan':
        return ''

    if dry_run:
        words = text.split()
        preview = ' '.join(words[:6]) + ('...' if len(words) > 6 else '')
        return f'[DRY-RUN: {len(words)} words -> rewrite "{preview}"]'

    intro = _FIELD_INTRO.get(field, 'שכתב את הטקסט הבא:')
    user_msg = f'{intro}\nשם הרב: {rabbi_name}\n\n{text}'

    resp = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=_SYSTEM,
        messages=[{'role': 'user', 'content': user_msg}],
    )
    return resp.content[0].text.strip()


# ── Checkpoint ────────────────────────────────────────────────────────────────

CHECKPOINT_FILE = os.path.join(os.path.dirname(__file__), 'enrich_checkpoint.json')


def load_checkpoint() -> dict:
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_checkpoint(data: dict):
    with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Excel helpers ─────────────────────────────────────────────────────────────

TARGET_FIELDS = ['biography', 'story', 'torah']

_FILL_ODD  = PatternFill('solid', fgColor='FFF8F0')
_FILL_EVEN = PatternFill('solid', fgColor='FFFFFF')
_FILL_HDR  = PatternFill('solid', fgColor='2C5F2E')
_FONT_HDR  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
_FONT_BODY = Font(name='Arial', size=10)
_ALIGN_RTL = Alignment(horizontal='right', vertical='top',
                        wrap_text=True, readingOrder=2)
_ALIGN_HDR = Alignment(horizontal='center', vertical='center',
                        readingOrder=2)

COL_WIDTHS = {
    'popular_name': 25, 'full_name': 30, 'birth_day': 18,
    'hilula_day': 10,   'hilula_day_num': 8, 'hilula_month': 12,
    'stream': 15,       'role': 15,          'image_filename': 25,
    'importance_score': 8, 'opening_quote': 40,
    'story': 60,        'torah': 60,
    'biography': 60,    'sources': 40,
}


def apply_sheet_style(ws):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = 'A2'

    # Header row
    for cell in ws[1]:
        cell.fill = _FILL_HDR
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_HDR

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = _FILL_ODD if row_idx % 2 else _FILL_EVEN
        for cell in row:
            cell.fill = fill
            cell.font = _FONT_BODY
            cell.alignment = _ALIGN_RTL

    # Column widths
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for col_idx, header in enumerate(headers, start=1):
        width = COL_WIDTHS.get(str(header), 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def save_excel(df: pd.DataFrame, excel_path: str, sheet_name: str):
    """Write df to the sheet, preserving other sheets in the workbook."""
    # Read existing workbook to preserve other sheets
    try:
        existing = pd.read_excel(excel_path, sheet_name=None, dtype=str)
    except Exception:
        existing = {}

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write the target sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Re-write other sheets
        for sname, sdf in existing.items():
            if sname != sheet_name:
                sdf.to_excel(writer, sheet_name=sname, index=False)

    # Apply styles
    wb = load_workbook(excel_path)
    for sname in wb.sheetnames:
        apply_sheet_style(wb[sname])
    wb.save(excel_path)
    print(f"  Saved {excel_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='AI enrichment for tzaddikim')
    parser.add_argument('--months', default='תשרי,חשוון',
                        help='Comma-separated Hebrew months (default: תשרי,חשוון)')
    parser.add_argument('--excel', default='tzaddikim_clean.xlsx')
    parser.add_argument('--sheet', default='tzaddikim')
    parser.add_argument('--dry-run', action='store_true',
                        help='Preview without calling API or saving')
    parser.add_argument('--resume', action='store_true',
                        help='Resume from checkpoint (skip already-processed rows)')
    parser.add_argument('--names-only', action='store_true',
                        help='Only normalize names, skip AI rewriting')
    args = parser.parse_args()

    months = [m.strip() for m in args.months.split(',')]
    print(f"Processing months: {months}", flush=True)
    print(f"Excel: {args.excel}  |  Dry-run: {args.dry_run}  |  Resume: {args.resume}", flush=True)

    # ── Load Excel ────────────────────────────────────────────────────────────
    print("Loading Excel...", flush=True)
    df = pd.read_excel(args.excel, sheet_name=args.sheet, dtype=str)
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns", flush=True)

    # ── Filter target rows ────────────────────────────────────────────────────
    mask = df['hilula_month'].isin(months)
    target_idx = df[mask].index.tolist()
    print(f"Target rows: {len(target_idx)}", flush=True)

    rows_with_content = [
        i for i in target_idx
        if any(
            str(df.loc[i, f]).strip() not in ('', 'nan')
            for f in TARGET_FIELDS
            if f in df.columns
        )
    ]
    print(f"  {len(rows_with_content)} have biography/story/torah content", flush=True)

    # ── Setup Claude ──────────────────────────────────────────────────────────
    client = None
    if not args.dry_run and not args.names_only:
        if not ANTHROPIC_KEY:
            print("ERROR: ANTHROPIC_API_KEY missing. Add to .env.local or environment.", flush=True)
            sys.exit(1)
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
            print(f"Claude model: {MODEL}", flush=True)
        except ImportError:
            print("ERROR: pip install anthropic", flush=True)
            sys.exit(1)

    # ── Load checkpoint ───────────────────────────────────────────────────────
    checkpoint = load_checkpoint() if args.resume else {}
    skipped_from_ckpt = 0

    # ── Process ───────────────────────────────────────────────────────────────
    processed = 0
    api_calls = 0
    errors = 0

    for idx in target_idx:
        row = df.loc[idx]
        ck_key = str(idx)

        # Apply checkpoint if resuming
        if ck_key in checkpoint:
            for field, value in checkpoint[ck_key].items():
                df.at[idx, field] = value
            skipped_from_ckpt += 1
            continue

        # --- Name normalization ---
        name_raw = str(row.get('popular_name', '') or '').strip()
        name_norm = normalize_name(name_raw) if name_raw and name_raw != 'nan' else name_raw
        df.at[idx, 'popular_name'] = name_norm
        row_ck = {'popular_name': name_norm}

        # --- AI rewriting (skipped if --names-only) ---
        if not args.names_only:
            for field in TARGET_FIELDS:
                if field not in df.columns:
                    continue
                text = str(row.get(field, '') or '').strip()
                if not text or text == 'nan':
                    continue

                try:
                    rewritten = rewrite_field(client, field, text, name_norm, args.dry_run)
                    df.at[idx, field] = rewritten
                    row_ck[field] = rewritten
                    api_calls += 1
                    if not args.dry_run:
                        time.sleep(DELAY_SEC)
                except Exception as e:
                    print(f"  ERROR row {idx} [{field}]: {e}")
                    errors += 1
                    row_ck[field] = text  # keep original on error

        # Save checkpoint after each row
        checkpoint[ck_key] = row_ck
        if not args.dry_run:
            save_checkpoint(checkpoint)

        processed += 1
        if processed % 10 == 0 or processed == 1:
            pct = 100 * processed // len(target_idx)
            print(f"  [{pct:3d}%] {processed}/{len(target_idx)} rows"
                  f" | {api_calls} API calls | {errors} errors", flush=True)

    # Final stats
    print(f"\nDone: {processed} processed, {skipped_from_ckpt} from checkpoint", flush=True)
    print(f"      {api_calls} API calls, {errors} errors", flush=True)

    # Sample output in dry-run
    if args.dry_run:
        sample = [i for i in target_idx[:3] if str(df.loc[i, 'biography']).strip() not in ('', 'nan')]
        for i in sample[:2]:
            print(f"\n  Row {i}: {df.loc[i, 'popular_name']}")
            bio = str(df.loc[i, 'biography']).strip()
            if bio and bio != 'nan':
                print(f"  biography: {bio[:120]}")

    # ── Save Excel ────────────────────────────────────────────────────────────
    if not args.dry_run:
        print(f"\nSaving {args.excel}...")
        save_excel(df, args.excel, args.sheet)
        print("Saved OK")
    else:
        print("\n[DRY-RUN] No files written.")


if __name__ == '__main__':
    main()
