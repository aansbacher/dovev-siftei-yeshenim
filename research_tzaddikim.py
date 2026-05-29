"""
research_tzaddikim.py
Web research agent: enriches tzaddikim from Wikipedia and hyomi.org.il.

For each tzaddik:
  1. Searches Wikipedia (Hebrew) for biography + image
  2. Fetches hyomi.org.il by hilula date for stories/quotes
  3. Fetches mytzadik.com by hilula date for additional info
  4. Uses Claude to synthesize: biography, story, torah, opening_quote
  5. Downloads image to public/images/tzaddikim/
  6. Updates tzaddikim_clean.xlsx + saves checkpoint

Usage:
    python research_tzaddikim.py --months סיון [--resume] [--overwrite] [--dry-run]

Month order: סיון → תמוז → אב → אלול → תשרי → חשוון → כסלו → טבת → שבט → אדר → ניסן → אייר

Requires:
    pip install anthropic requests beautifulsoup4 openpyxl pandas
    ANTHROPIC_API_KEY in .env.local
"""

import os, re, sys, io, time, json, argparse, urllib.parse
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Paths & Config ────────────────────────────────────────────────────────────

BASE_DIR     = Path(__file__).parent
IMAGES_DIR   = BASE_DIR / 'public' / 'images' / 'tzaddikim'
CHECKPOINT_F = BASE_DIR / 'research_checkpoint.json'
EXCEL_PATH   = BASE_DIR / 'tzaddikim_clean.xlsx'

MODEL      = 'claude-haiku-4-5-20251001'
DELAY_SEC  = 2.0   # between web requests

# Hebrew month → number (from Tishri=1)
HEBREW_MONTH_NUM = {
    'תשרי': 1, 'חשוון': 2, 'חשון': 2, 'כסלו': 3, 'טבת': 4,
    'שבט': 5, 'אדר': 6, 'אדר א': 6, 'אדר ב': 7,
    'ניסן': 8, 'אייר': 9, 'סיון': 10, 'תמוז': 11, 'אב': 12, 'אלול': 13,
}

_HONORIFICS = [
    'מרן הרב', 'מרן רבי', 'הגאון הגדול הרב', 'הגאון הגדול רבי',
    'הגאון הרב', 'הגאון רבי', 'הרה"ג', 'הרה"צ', 'הרה"ק',
    'האדמו"ר', 'מרן', 'הרב', 'רבנו', 'רבן', 'רבי', 'חכם',
]
_SUFFIX_RE       = re.compile(
    r'\s*,?\s*(?:זצ"ל|זיע"א|זי"ע|זצוק"ל|ז"ל|נ"ע|עה"ש'
    r'|זכרונו לברכה|זכר צדיק לברכה)\s*$'
)
_DATE_PARENS_RE  = re.compile(r'\s*\([^)]*(?:\d{3,4}|ת[רשק]["א-ת]+)[^)]*\)')
_NIKUD_RE        = re.compile(r'[ְ-ׇ]')

LETTER_VAL = {
    'א':1,'ב':2,'ג':3,'ד':4,'ה':5,'ו':6,'ז':7,'ח':8,'ט':9,
    'י':10,'כ':20,'ל':30,
}

# ── Env ───────────────────────────────────────────────────────────────────────

def load_env(path='.env.local'):
    env = {}
    fp = os.path.join(os.path.dirname(__file__), path)
    if os.path.exists(fp):
        with open(fp, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1)
                    env[k.strip()] = v.strip()
    return env

_env = load_env()
ANTHROPIC_KEY = os.getenv('ANTHROPIC_API_KEY') or _env.get('ANTHROPIC_API_KEY', '')

# ── Name helpers ──────────────────────────────────────────────────────────────

def strip_honorifics(name: str) -> str:
    clean = _NIKUD_RE.sub('', str(name).strip())
    for h in sorted(_HONORIFICS, key=len, reverse=True):
        if clean.startswith(h):
            clean = clean[len(h):].strip()
            break
    clean = _SUFFIX_RE.sub('', clean).strip().strip(',').strip()
    clean = _DATE_PARENS_RE.sub('', clean).strip()
    return clean


def clean_for_search(name: str) -> str:
    """Best search term: strip honorifics/suffix/dates from full_name."""
    if not name or str(name).strip() in ('', 'nan'):
        return ''
    return strip_honorifics(name)


def safe_filename(name: str) -> str:
    """Filesystem-safe name for image files."""
    clean = strip_honorifics(str(name))
    clean = re.sub(r'[^א-ת ]', '', clean).strip()
    clean = re.sub(r'\s+', '_', clean)
    return clean[:80] if clean else 'tzaddik'


def day_to_int(day_str: str) -> int | None:
    clean = re.sub(r"""['״"״\s]""", '', str(day_str or '')).strip()
    val = sum(LETTER_VAL.get(c, 0) for c in clean)
    return val if 1 <= val <= 30 else None


def val_str(s) -> str:
    v = str(s or '').strip()
    return '' if v == 'nan' else v

# ── HTTP session ──────────────────────────────────────────────────────────────

_http = requests.Session()
_http.headers.update({
    'User-Agent': 'TzaddikimResearchBot/1.0 (educational; contact: admin@example.com)',
    'Accept-Language': 'he,en;q=0.9',
})

# ── Wikipedia ─────────────────────────────────────────────────────────────────

_WIKI_SEARCH = (
    'https://he.wikipedia.org/w/api.php?action=query&list=search'
    '&srsearch={q}&format=json&utf8=1&srlimit=5&srnamespace=0'
)
_WIKI_FETCH = (
    'https://he.wikipedia.org/w/api.php?action=query'
    '&prop=extracts|pageimages&exintro=1&explaintext=1'
    '&piprop=original|thumbnail&pithumbsize=600&titles={title}&format=json&utf8=1'
)


def search_wikipedia(name: str) -> str | None:
    """Return best Wikipedia title for a name, or None."""
    if not name:
        return None
    q = urllib.parse.quote(name)
    try:
        r = _http.get(_WIKI_SEARCH.format(q=q), timeout=10)
        hits = r.json().get('query', {}).get('search', [])
        if hits:
            return hits[0]['title']
    except Exception:
        pass
    return None


def fetch_wikipedia(title: str) -> dict:
    """Return {'extract': str, 'image_url': str|None}."""
    enc = urllib.parse.quote(title)
    try:
        r = _http.get(_WIKI_FETCH.format(title=enc), timeout=10)
        pages = r.json().get('query', {}).get('pages', {})
        page  = next(iter(pages.values()))
        if page.get('pageid', -1) == -1:
            return {'extract': '', 'image_url': None}
        extract   = (page.get('extract') or '')[:1500]
        image_url = (
            (page.get('original') or {}).get('source') or
            (page.get('thumbnail') or {}).get('source')
        )
        return {'extract': extract, 'image_url': image_url}
    except Exception:
        return {'extract': '', 'image_url': None}

# ── hyomi.org.il ─────────────────────────────────────────────────────────────

_HYOMI_URL = 'https://hyomi.org.il/day.php?month1=0&dd={day}&mm={month}&yy=2026'


def fetch_hyomi(day_num: int, month: str) -> str:
    """Fetch hyomi.org.il for a hilula date, return page text."""
    month_num = HEBREW_MONTH_NUM.get(month)
    if not month_num or not day_num:
        return ''
    url = _HYOMI_URL.format(day=day_num, month=month_num)
    try:
        r = _http.get(url, timeout=12)
        soup = BeautifulSoup(r.text, 'html.parser')
        for tag in soup(['script', 'style', 'nav', 'footer']):
            tag.decompose()
        return soup.get_text(separator='\n', strip=True)[:5000]
    except Exception:
        return ''


def extract_from_hyomi(page_text: str, search_name: str) -> str:
    """Find the paragraph(s) about a specific tzaddik in a hyomi page."""
    if not page_text or not search_name:
        return ''
    # Use significant words from name (skip common prefixes)
    words = [w for w in search_name.split() if len(w) >= 3]
    for word in words:
        idx = page_text.find(word)
        if idx >= 0:
            start = max(0, idx - 100)
            end   = min(len(page_text), idx + 1200)
            return page_text[start:end]
    return ''

# ── mytzadik.com ─────────────────────────────────────────────────────────────

_MYTZADIK_SEARCH = 'https://www.mytzadik.com/search.aspx?q={q}'


def fetch_mytzadik(name: str) -> str:
    """Try to search mytzadik.com for the name, return page text."""
    q = urllib.parse.quote(name)
    try:
        r = _http.get(_MYTZADIK_SEARCH.format(q=q), timeout=10)
        soup = BeautifulSoup(r.text, 'html.parser')
        for tag in soup(['script', 'style']):
            tag.decompose()
        return soup.get_text(separator='\n', strip=True)[:3000]
    except Exception:
        return ''

# ── Image download ────────────────────────────────────────────────────────────

def download_image(url: str, name: str) -> str | None:
    """Download image to IMAGES_DIR, return relative filename or None."""
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    ext = url.rsplit('.', 1)[-1].split('?')[0].lower()
    if ext not in ('jpg', 'jpeg', 'png', 'webp'):
        ext = 'jpg'
    fname    = f'{safe_filename(name)}.{ext}'
    filepath = IMAGES_DIR / fname
    if filepath.exists():
        return fname
    for attempt in range(4):
        try:
            r = _http.get(url, timeout=30, stream=True)
            if r.status_code == 200:
                with open(filepath, 'wb') as f:
                    for chunk in r.iter_content(8192):
                        f.write(chunk)
                return fname
            elif r.status_code == 429:
                wait = int(r.headers.get('Retry-After', 65))
                print(f'    [image 429] waiting {wait}s...', flush=True)
                time.sleep(wait + 5)
            else:
                break
        except Exception:
            time.sleep(5)
    return None

# ── Claude synthesis ──────────────────────────────────────────────────────────

_SYSTEM = """\
אתה עורך תוכן רוחני לאפליקציה יהודית-חסידית על צדיקים ורבנים.
הסגנון: חם, מכובד, נוגע ללב — כפי שמספרים על צדיקים בבית מדרש.

כללים:
- עברית פשוטה וברורה, ללא ניקוד.
- לא יותר מ-120 מילים לכל שדה.
- שמור על עובדות מהמקורות בלבד — אל תמציא.
- כתוב בגוף שלישי.
- אל תוסיף כותרת — רק הטקסט עצמו.
- אם אין מידע מספיק — החזר מחרוזת ריקה.
"""

_SYSTEM_STORIES = """\
אתה עורך תוכן רוחני לאפליקציה יהודית-חסידית על צדיקים ורבנים.
הסגנון: חם, מכובד, נוגע ללב — כפי שמספרים על צדיקים בבית מדרש.

כללים:
- עברית פשוטה וברורה, ללא ניקוד.
- לא יותר מ-120 מילים לכל שדה.
- כתוב בגוף שלישי.
- אל תוסיף כותרת — רק הטקסט עצמו.
- לסיפור: כתוב סיפור קצר ברוח חייו של הרב ("מסופר כי..."). אפשר לבסס על העובדות הביוגרפיות.
- לדבר תורה: כתוב אמרה או עיקרון רוחני המשקף את דרכו ותורתו.
- לציטוט: משפט אחד כאמרה ישירה בסגנון הרב.
- אם אין שום מידע לעבוד איתו — החזר מחרוזת ריקה.
"""

_PROMPT_FULL = """\
להלן מידע שנמצא ברשת על {name}:

{sources}

בהתבסס אך ורק על המידע שלמעלה, כתוב:

<biography>ביוגרפיה עד 120 מילים, גוף שלישי. ריק אם אין מידע.</biography>
<story>סיפור אחד מעניין, עד 120 מילים. ריק אם אין.</story>
<torah>דבר תורה או אמרה המיוחסת לו. ריק אם אין.</torah>
<opening_quote>משפט אחד קצר, ציטוט ישיר. ריק אם אין.</opening_quote>

כתוב רק את ארבע התגיות, ללא הסברים נוספים.
"""

_PROMPT_NO_BIO = """\
להלן מידע שנמצא ברשת על {name}:

{sources}

ביוגרפיה כבר קיימת — כתוב רק:

<story>סיפור אחד מעניין, עד 120 מילים. ריק אם אין.</story>
<torah>דבר תורה או אמרה המיוחסת לו. ריק אם אין.</torah>
<opening_quote>משפט אחד קצר, ציטוט ישיר. ריק אם אין.</opening_quote>

כתוב רק את שלוש התגיות, ללא הסברים נוספים.
"""

_PROMPT_FILL_STORIES = """\
להלן ביוגרפיה של {name}:

{biography}

בהתבסס על המידע לעיל, כתוב:

<story>סיפור קצר ברוח חייו (מסופר כי...), עד 100 מילים. אם אין שום מידע — ריק.</story>
<torah>אמרה רוחנית או עיקרון המשקף את דרכו ותורתו, עד 60 מילים. אם אין — ריק.</torah>
<opening_quote>משפט אחד קצר כציטוט ישיר בסגנון הרב. אם אין — ריק.</opening_quote>

כתוב רק את שלוש התגיות, ללא הסברים.
"""


def parse_response(text: str) -> dict:
    result = {}
    for field in ('biography', 'story', 'torah', 'opening_quote'):
        m = re.search(rf'<{field}>(.*?)</{field}>', text, re.DOTALL)
        if m:
            val = m.group(1).strip()
            if val:
                result[field] = val
    return result


def synthesize(client, name: str, sources: dict) -> dict:
    parts = []
    if sources.get('wikipedia'):
        parts.append(f'=== ויקיפדיה ===\n{sources["wikipedia"]}')
    if sources.get('hyomi'):
        parts.append(f'=== hyomi.org.il ===\n{sources["hyomi"]}')
    if sources.get('mytzadik'):
        parts.append(f'=== mytzadik.com ===\n{sources["mytzadik"]}')

    if not parts:
        return {}

    bio_exists = sources.get('_bio_exists', False)
    template = _PROMPT_NO_BIO if bio_exists else _PROMPT_FULL
    prompt = template.format(name=name, sources='\n\n'.join(parts))
    try:
        resp = client.messages.create(
            model=MODEL,
            max_tokens=900,
            system=_SYSTEM,
            messages=[{'role': 'user', 'content': prompt}],
        )
        return parse_response(resp.content[0].text.strip())
    except Exception as e:
        print(f'    [Claude] {e}', flush=True)
    return {}

# ── Checkpoint ────────────────────────────────────────────────────────────────

def load_checkpoint() -> dict:
    if CHECKPOINT_F.exists():
        with open(CHECKPOINT_F, encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_checkpoint(ck: dict):
    with open(CHECKPOINT_F, 'w', encoding='utf-8') as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)

# ── Excel save ────────────────────────────────────────────────────────────────

def save_excel(df: pd.DataFrame, path: str, sheet: str):
    try:
        existing = pd.read_excel(path, sheet_name=None, dtype=str)
    except Exception:
        existing = {}
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)
        for s, d in existing.items():
            if s != sheet:
                d.to_excel(writer, sheet_name=s, index=False)
    print(f'  Saved {path}', flush=True)

# ── Images-only pass ─────────────────────────────────────────────────────────

def _run_images_only(args):
    """Download images for all rows in checkpoint that have _wiki_title but no image file."""
    checkpoint = load_checkpoint()
    df = pd.read_excel(args.excel, sheet_name=args.sheet, dtype=str)

    months = [m.strip() for m in args.months.split(',')]
    mask = df['hilula_month'].isin(months)
    target_idx = df[mask].index.tolist()

    todo = []
    for idx in target_idx:
        ck = checkpoint.get(str(idx), {})
        wiki_title = ck.get('_wiki_title')
        if not wiki_title:
            continue
        img_val  = val_str(df.loc[idx].get('image_filename'))
        img_real = bool(img_val) and (IMAGES_DIR / img_val).exists()
        if img_real:
            continue
        todo.append((idx, ck.get('_name', ''), wiki_title))

    print(f'Images-only: {len(todo)} rows need images', flush=True)
    downloaded = 0

    for i, (idx, name, wiki_title) in enumerate(todo):
        wiki_data = fetch_wikipedia(wiki_title)
        time.sleep(DELAY_SEC)
        if wiki_data.get('image_url'):
            fname = download_image(wiki_data['image_url'], name)
            if fname:
                df.at[idx, 'image_filename'] = fname
                checkpoint[str(idx)]['image_filename'] = fname
                downloaded += 1
                save_checkpoint(checkpoint)
                print(f'  [{i+1}/{len(todo)}] {name} → {fname}', flush=True)
            else:
                print(f'  [{i+1}/{len(todo)}] {name} — download failed', flush=True)
        else:
            print(f'  [{i+1}/{len(todo)}] {name} — no image on Wikipedia', flush=True)
        time.sleep(3)  # extra delay between image downloads

    print(f'\nDone: {downloaded}/{len(todo)} images downloaded')
    if downloaded > 0:
        save_excel(df, args.excel, args.sheet)


# ── Fill-stories pass ─────────────────────────────────────────────────────────

def _run_fill_stories(args):
    """For every row missing story/torah, generate them from existing biography using Claude."""
    if not ANTHROPIC_KEY:
        print('ERROR: ANTHROPIC_API_KEY missing in .env.local'); sys.exit(1)
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    print(f'Claude: {MODEL}', flush=True)

    df = pd.read_excel(args.excel, sheet_name=args.sheet, dtype=str)
    months_filter = [m.strip() for m in args.months.split(',')]
    if months_filter != ['all']:
        mask = df['hilula_month'].isin(months_filter)
        target_idx = df[mask].index.tolist()
    else:
        target_idx = df.index.tolist()

    checkpoint = load_checkpoint()

    todo = []
    for idx in target_idx:
        row = df.loc[idx]
        bio   = val_str(row.get('biography'))
        story = val_str(row.get('story'))
        torah = val_str(row.get('torah'))
        if bio and (not story or not torah):
            todo.append(idx)

    print(f'Fill-stories: {len(todo)} rows missing story/torah (have bio)', flush=True)
    updated = api_calls = 0

    for i, idx in enumerate(todo):
        row  = df.loc[idx]
        name = val_str(row.get('popular_name')) or val_str(row.get('full_name')) or '?'
        bio  = val_str(row.get('biography'))

        pct = int((i + 1) / len(todo) * 100)
        if (i + 1) % 5 == 0 or i == 0:
            print(f'  [{pct:3d}%] {i+1}/{len(todo)} | updated={updated}', flush=True)

        prompt = _PROMPT_FILL_STORIES.format(name=name, biography=bio)
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=600,
                system=_SYSTEM_STORIES,
                messages=[{'role': 'user', 'content': prompt}],
            )
            result = parse_response(resp.content[0].text.strip())
            api_calls += 1
            time.sleep(0.5)
        except Exception as e:
            print(f'    [Claude] {e}', flush=True)
            time.sleep(5)
            continue

        changed = False
        ck_entry = checkpoint.get(str(idx), {})
        for field in ('story', 'torah', 'opening_quote'):
            existing = val_str(row.get(field))
            if result.get(field) and result[field].strip() and not existing:
                df.at[idx, field] = result[field]
                ck_entry[field] = result[field]
                changed = True
        if changed:
            checkpoint[str(idx)] = ck_entry
            save_checkpoint(checkpoint)
            updated += 1

        time.sleep(DELAY_SEC)

    print(f'\nSaving Excel...', flush=True)
    save_excel(df, args.excel, args.sheet)
    print(f'Done: {updated} rows updated | Claude calls: {api_calls}', flush=True)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Web research for tzaddikim')
    parser.add_argument('--months',    default='סיון',
                        help='Comma-separated Hebrew months, or "all" (default: סיון)')
    parser.add_argument('--excel',     default=str(EXCEL_PATH))
    parser.add_argument('--sheet',     default='tzaddikim')
    parser.add_argument('--resume',    action='store_true',
                        help='Skip rows already in checkpoint')
    parser.add_argument('--overwrite', action='store_true',
                        help='Overwrite existing biography/story/torah')
    parser.add_argument('--dry-run',    action='store_true',
                        help='Search and print only, do not save')
    parser.add_argument('--images-only', action='store_true',
                        help='Only download images for rows in checkpoint (no text research)')
    parser.add_argument('--skip-images', action='store_true',
                        help='Skip image download during research (do it later with --images-only)')
    parser.add_argument('--fill-stories', action='store_true',
                        help='For rows with bio but no story/torah: generate from existing bio')
    args = parser.parse_args()

    months = [m.strip() for m in args.months.split(',')]
    print(f'Months: {months} | Resume: {args.resume} | Overwrite: {args.overwrite}', flush=True)

    # Images-only mode: download images for checkpoint rows that have a wiki title but no image
    if args.images_only:
        _run_images_only(args)
        return

    # Fill-stories mode: generate story/torah/quote from existing biography
    if args.fill_stories:
        _run_fill_stories(args)
        return

    # Setup Claude
    client = None
    if not args.dry_run:
        if not ANTHROPIC_KEY:
            print('ERROR: ANTHROPIC_API_KEY missing in .env.local'); sys.exit(1)
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        print(f'Claude: {MODEL}', flush=True)

    # Load Excel
    df = pd.read_excel(args.excel, sheet_name=args.sheet, dtype=str)
    print(f'Loaded {len(df)} rows, {len(df.columns)} columns', flush=True)

    mask       = df['hilula_month'].isin(months)
    target_idx = df[mask].index.tolist()
    print(f'Target: {len(target_idx)} rows', flush=True)

    checkpoint  = load_checkpoint() if args.resume else {}
    processed   = api_calls = images_found = wiki_found = 0

    for i, idx in enumerate(target_idx):
        row    = df.loc[idx]
        ck_key = str(idx)

        # Resume: restore from checkpoint and skip
        if args.resume and ck_key in checkpoint:
            saved = checkpoint[ck_key]
            for field, value in saved.items():
                if not field.startswith('_') and field in df.columns:
                    df.at[idx, field] = value
            continue

        name      = val_str(row.get('popular_name'))
        full_name = val_str(row.get('full_name'))
        search    = clean_for_search(full_name or name)

        # What already exists?
        img_val  = val_str(row.get('image_filename'))
        img_real = bool(img_val) and (IMAGES_DIR / img_val).exists()
        has = {
            'biography':      bool(val_str(row.get('biography'))),
            'story':          bool(val_str(row.get('story'))),
            'torah':          bool(val_str(row.get('torah'))),
            'opening_quote':  bool(val_str(row.get('opening_quote'))),
            'image_filename': img_real,
        }

        needs = (
            args.overwrite or
            not has['biography'] or not has['story'] or not has['torah']
        )

        row_ck = {'_name': name}

        if needs and not args.dry_run:
            sources = {}

            # 1. Wikipedia
            if search:
                wiki_title = search_wikipedia(search)
                time.sleep(DELAY_SEC)
                if wiki_title:
                    wiki_data = fetch_wikipedia(wiki_title)
                    time.sleep(DELAY_SEC)
                    if wiki_data['extract']:
                        sources['wikipedia'] = wiki_data['extract']
                        row_ck['_wiki_title'] = wiki_title
                        wiki_found += 1

                    if not has['image_filename'] and wiki_data.get('image_url') and not args.skip_images:
                        fname = download_image(wiki_data['image_url'], name)
                        if fname:
                            df.at[idx, 'image_filename'] = fname
                            row_ck['image_filename'] = fname
                            images_found += 1
                    elif wiki_data.get('image_url'):
                        row_ck['_image_url'] = wiki_data['image_url']

            # 2. hyomi.org.il
            day_num = day_to_int(val_str(row.get('hilula_day')))
            month   = val_str(row.get('hilula_month'))
            if day_num and month:
                hyomi_page = fetch_hyomi(day_num, month)
                time.sleep(DELAY_SEC)
                snippet = extract_from_hyomi(hyomi_page, search)
                if snippet:
                    sources['hyomi'] = snippet

            # (mytzadik.com removed — site blocks bots; dirshu.co.il also blocks bots)

            # Mark if biography already exists (so Claude skips it)
            if has['biography']:
                sources['_bio_exists'] = True

            # 4. Claude synthesis
            if sources:
                result = synthesize(client, name, sources)
                api_calls += 1
                time.sleep(0.5)

                for field in ('biography', 'story', 'torah', 'opening_quote'):
                    if result.get(field) and result[field].strip():
                        if not has[field] or args.overwrite:
                            df.at[idx, field] = result[field]
                            row_ck[field] = result[field]

        elif args.dry_run and needs:
            print(f'  [DRY] {name} | search: {search}', flush=True)

        # Save checkpoint after every row
        checkpoint[ck_key] = row_ck
        if not args.dry_run:
            save_checkpoint(checkpoint)

        processed += 1

        if processed % 5 == 0 or processed == 1:
            pct = 100 * (i + 1) // len(target_idx)
            print(
                f'  [{pct:3d}%] {i+1}/{len(target_idx)} | '
                f'Wiki: {wiki_found} | API: {api_calls} | Images: {images_found}',
                flush=True
            )

    # Save Excel
    if not args.dry_run:
        print(f'\nSaving Excel...', flush=True)
        save_excel(df, args.excel, args.sheet)

    print(
        f'\nDone: {processed} rows | '
        f'Wikipedia found: {wiki_found} | '
        f'Claude calls: {api_calls} | '
        f'Images: {images_found}',
        flush=True
    )


if __name__ == '__main__':
    main()
